# pyright: reportAttributeAccessIssue=false
"""Excel parsing mixin for AttachmentParser.

Handles Excel (.xlsx) text and table extraction via openpyxl.
"""

from __future__ import annotations

from pathlib import Path

from .models import AttachmentParseResult
from ._attachment_helpers import _text_chars


class _ExcelParserMixin:
    """Excel parsing methods for AttachmentParser."""

    @staticmethod
    def _process_excel_sheet(
        sheet, sheet_name: str,
    ) -> tuple[list[dict], list[str]]:
        """Extract table data and text lines from a single Excel sheet.

        Returns (tables, text_parts).
        """
        tables: list[dict] = []
        text_parts: list[str] = []

        rows_data = []
        for row in sheet.iter_rows(values_only=True):
            row_values = [
                str(cell) if cell is not None else "" for cell in row
            ]
            if any(v.strip() for v in row_values):
                rows_data.append(row_values)

        if not rows_data:
            return tables, text_parts

        headers = rows_data[0]
        data_rows = rows_data[1:] if len(rows_data) > 1 else []

        tables.append({
            "sheet": sheet_name,
            "headers": headers,
            "rows": [
                dict(zip(headers, row))
                for row in data_rows
                if len(row) == len(headers)
            ],
            "row_count": len(data_rows),
        })

        text_parts.append(f"[Sheet: {sheet_name}]")
        text_parts.append(" | ".join(headers))
        for row in data_rows[:10]:
            text_parts.append(" | ".join(row))
        if len(data_rows) > 10:
            text_parts.append(f"... 외 {len(data_rows) - 10}행")

        return tables, text_parts

    @classmethod
    def parse_excel(cls, file_path: Path) -> AttachmentParseResult:
        """Excel에서 시트 데이터 추출"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            text_parts: list[str] = []
            tables: list[dict] = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_tables, sheet_texts = (
                    cls._process_excel_sheet(
                        sheet, sheet_name,
                    )
                )
                tables.extend(sheet_tables)
                text_parts.extend(sheet_texts)

            wb.close()

            full_text = "\n".join(text_parts)
            confidence = 0.95 if tables else 0.0

            return AttachmentParseResult(
                extracted_text=full_text,
                extracted_tables=tables,
                confidence=confidence,
                native_text_chars=_text_chars(full_text),
            )

        except (
            RuntimeError, OSError, ValueError,
            TypeError, KeyError, AttributeError,
        ) as e:
            return AttachmentParseResult(
                extracted_text=f"[Excel 파싱 오류: {e}]",
                extracted_tables=[],
                confidence=0.0,
                ocr_skip_reason="parse_error",
            )
